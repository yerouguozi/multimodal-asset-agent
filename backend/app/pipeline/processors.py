"""模态适配器：把"理解素材"这件事按模态拆成可插拔的处理器。

新增模态 = 新增一个 process_xxx 函数并注册，核心引擎不变。
阶段 7 增强：
- 模型路由：小图走轻量模型（8B），大图走大模型（32B），控制成本；
- 每次模型调用记录 UsageLog（成本追踪）。
"""
from __future__ import annotations

import base64
import io
import json
import logging
import math
import re
import shutil
import subprocess
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

from PIL import Image, ImageDraw

import imagehash

from ..core.config import settings
from ..llm.client import MultimodalClient
from ..models import Asset
from ..retrieval.vector_store import vector_store
from ..usage import record_usage

logger = logging.getLogger(__name__)

MAX_VISION_SIDE = 1280  # 发给视觉模型前的最大边长（控制 token/费用）
MAX_THUMB_SIDE = 320
TEXT_LIMIT = 5000  # 入库保存的正文上限（字符）


@dataclass
class ProcessingResult:
    description: str | None = None
    tags: list[str] = field(default_factory=list)
    ocr_text: str | None = None
    transcript: str | None = None
    transcript_segments: str | None = None
    text_content: str | None = None
    thumbnail_path: str | None = None
    phash: str | None = None
    width: int | None = None
    height: int | None = None
    duration: float | None = None
    vision_model: str | None = None


# ---------- 通用工具 ----------

def _ffmpeg() -> str:
    import imageio_ffmpeg

    return imageio_ffmpeg.get_ffmpeg_exe()


def _run(cmd: list[str], timeout: int = 120) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)


def _image_to_b64(img: Image.Image, max_side: int = MAX_VISION_SIDE) -> str:
    probe = img.copy()
    probe.thumbnail((max_side, max_side))
    buf = io.BytesIO()
    probe.convert("RGB").save(buf, "JPEG", quality=85)
    return base64.b64encode(buf.getvalue()).decode("ascii")


def _media_duration(path: Path) -> float | None:
    """用 ffmpeg -i 的 stderr 解析时长（ffprobe 未随包分发）。"""
    try:
        r = _run([_ffmpeg(), "-i", str(path)])
        m = re.search(r"Duration:\s*(\d+):(\d+):(\d+(?:\.\d+)?)", r.stderr or "")
        if m:
            return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + float(m.group(3))
    except Exception as e:
        logger.warning("解析时长失败 %s: %s", path.name, e)
    return None


def _route_vision_model(max_side: int) -> str:
    """模型路由：小图走轻量模型，大图走大模型（成本优化的核心规则）。"""
    if max_side <= settings.simple_image_max_side:
        return settings.vision_model_cheap
    return settings.vision_model


def _extract_keyframes(src: Path, out_dir: Path, max_frames: int, duration: float | None) -> list[Path]:
    """均匀抽帧：按时长把 max_frames 张关键帧均匀铺开。"""
    out_dir.mkdir(parents=True, exist_ok=True)
    vf = f"fps=1/{max(0.1, duration / max_frames):.2f}" if duration else "fps=1/2"
    _run([_ffmpeg(), "-y", "-i", str(src), "-vf", vf, "-frames:v", str(max_frames), str(out_dir / "f_%02d.jpg")])
    return sorted(out_dir.glob("f_*.jpg"))


def _slice_audio(src: Path, start: float, duration: float, out: Path) -> Path | None:
    """切出 [start, start+duration) 的 16k 单声道 wav。"""
    out.parent.mkdir(parents=True, exist_ok=True)
    _run([_ffmpeg(), "-y", "-ss", f"{start:.2f}", "-t", f"{duration:.2f}", "-i", str(src),
          "-acodec", "pcm_s16le", "-ar", "16000", "-ac", "1", str(out)])
    return out if out.exists() and out.stat().st_size > 0 else None


def _transcribe_segments(asset: Asset, llm: MultimodalClient, wav_path: Path, work_dir: Path, duration: float | None):
    """整段或分片转写。返回 (全文, segments[{start,end,text}])，供"找说过某段话"使用。

    分片数按提取窗口（min(duration, audio_max_seconds)，wav 本身就截到这里）计算，
    并受 max_asr_chunks 上限约束——上限必须覆盖整个窗口，否则尾部内容搜不到。
    只对真实发生的转写调用记账。
    """
    chunk = settings.asr_chunk_seconds
    if not duration or duration <= chunk:
        text = llm.transcribe_audio(wav_path)
        if not text:
            return None, None
        text = text.strip()
        record_usage(asset.id, settings.asr_model, "asr")
        return text, [{"start": 0, "end": round(duration or 0, 1), "text": text}]

    window = min(duration, settings.audio_max_seconds)
    n = min(int(math.ceil(window / chunk)), settings.max_asr_chunks)
    segments: list[dict] = []
    parts: list[str] = []
    for i in range(n):
        start = i * chunk
        seg = _slice_audio(wav_path, start, chunk, work_dir / f"seg_{i}.wav")
        if not seg:
            continue
        text = llm.transcribe_audio(seg)
        if text and text.strip():
            text = text.strip()
            record_usage(asset.id, settings.asr_model, "asr")
            parts.append(text)
            segments.append({
                "start": round(start, 1),
                "end": round(min(start + chunk, duration or start + chunk), 1),
                "text": text,
            })
    return (" ".join(parts)) or None, segments

def _extract_audio(src: Path, out_wav: Path, max_seconds: int | None) -> Path | None:
    """抽音轨转 16kHz 单声道 wav（ASR 标准输入）；可限制时长控制成本。"""
    out_wav.parent.mkdir(parents=True, exist_ok=True)
    cmd = [_ffmpeg(), "-y", "-i", str(src), "-vn", "-acodec", "pcm_s16le", "-ar", "16000", "-ac", "1"]
    if max_seconds:
        cmd += ["-t", str(max_seconds)]
    cmd.append(str(out_wav))
    _run(cmd)
    return out_wav if out_wav.exists() and out_wav.stat().st_size > 0 else None


def _first_frame(src: Path, thumb_path: Path) -> Path | None:
    """视频封面：优先取第 1 秒，失败退回第 0 帧。"""
    thumb_path.parent.mkdir(parents=True, exist_ok=True)
    _run([_ffmpeg(), "-y", "-ss", "1", "-i", str(src), "-frames:v", "1", "-q:v", "3", str(thumb_path)])
    if not thumb_path.exists():
        _run([_ffmpeg(), "-y", "-i", str(src), "-frames:v", "1", "-q:v", "3", str(thumb_path)])
    return thumb_path if thumb_path.exists() else None


def _audio_placeholder(thumb_path: Path) -> Path:
    """音频没有画面，生成一个带 AUDIO 字样的占位缩略图。"""
    thumb_path.parent.mkdir(parents=True, exist_ok=True)
    img = Image.new("RGB", (320, 180), (30, 41, 59))
    draw = ImageDraw.Draw(img)
    text = "AUDIO"
    try:
        bbox = draw.textbbox((0, 0), text)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    except Exception:
        tw, th = 80, 20
    draw.text(((320 - tw) / 2, (180 - th) / 2), text, fill=(148, 163, 184))
    img.save(thumb_path, "JPEG", quality=80)
    return thumb_path


def _top_tags(tags: list[str], limit: int = 10) -> list[str]:
    return [t for t, _ in Counter(tags).most_common(limit)]


def _dedupe_join(items: list[str], sep: str = "；", limit: int = 1000) -> str | None:
    merged = sep.join(dict.fromkeys(x for x in items if x))
    return merged[:limit] or None


# ---------- 图片 ----------

def process_image(asset: Asset, llm: MultimodalClient, data_root: Path) -> ProcessingResult:
    path = data_root / asset.storage_path
    result = ProcessingResult()

    with Image.open(path) as img:
        img.load()
        result.width, result.height = img.size
        result.phash = str(imagehash.phash(img, hash_size=16))

        thumb_dir = data_root / "thumbnails"
        thumb_dir.mkdir(parents=True, exist_ok=True)
        thumb = img.copy()
        thumb.thumbnail((MAX_THUMB_SIDE, MAX_THUMB_SIDE))
        thumb_path = thumb_dir / f"{asset.id}.jpg"
        thumb.convert("RGB").save(thumb_path, "JPEG", quality=80)
        result.thumbnail_path = f"thumbnails/{thumb_path.name}"

        b64 = _image_to_b64(img)
        max_side = max(img.size)

    model = _route_vision_model(max_side)
    result.vision_model = model
    vision = llm.vision_describe(b64, "image/jpeg", model=model)
    if vision:
        record_usage(asset.id, model, "vision")
        result.description = vision.description
        result.tags = vision.tags
        result.ocr_text = vision.ocr

    # 多模态 embedding：图片直接嵌入（检索时与文本向量做三路融合）
    try:
        vl_vec = llm.embed_image(b64)
        if vl_vec:
            vector_store.add(asset.id, vl_vec, settings.vl_embedding_model)
            record_usage(asset.id, settings.vl_embedding_model, "vl_embed")
    except Exception as e:
        logger.warning("VL 图片嵌入失败（已降级）: %s", e)
    return result


# ---------- 文档 ----------

def process_document(asset: Asset, llm: MultimodalClient, data_root: Path) -> ProcessingResult:
    path = data_root / asset.storage_path
    result = ProcessingResult()
    text = _extract_text(path, asset.mime_type)

    if text:
        result.text_content = text[:TEXT_LIMIT]
        summary = llm.summarize_text(text[:3000])
        if summary:
            record_usage(asset.id, settings.llm_model, "summary")
            result.description = summary.summary
            result.tags = summary.tags
    else:
        result.description = "（未提取到文本，可能是扫描件；OCR 支持将在后续阶段补充）"
    return result


def _extract_text(path: Path, mime: str) -> str:
    suffix = path.suffix.lower()
    try:
        if suffix == ".pdf":
            return _extract_pdf(path)
        if suffix in (".docx", ".doc"):
            return _extract_docx(path)
        if suffix in (".xlsx", ".xlsm", ".xls"):
            return _extract_xlsx(path)
        if suffix in (".txt", ".md", ".json", ".csv", ".log"):
            data = path.read_bytes()
            return data[:200_000].decode("utf-8", errors="ignore")
    except Exception as e:
        logger.warning("文档文本抽取失败 %s: %s", path.name, e)
        return ""
    return ""


def _extract_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts = []
    for page in reader.pages[:50]:
        parts.append(page.extract_text() or "")
    return "\n".join(parts).strip()


def _extract_docx(path: Path) -> str:
    from docx import Document

    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip()).strip()


def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets[:3]:
        for row in ws.iter_rows(max_row=50, values_only=True):
            vals = [str(v) for v in row if v is not None]
            if vals:
                lines.append(" | ".join(vals))
    return "\n".join(lines).strip()


# ---------- 视频 ----------

def process_video(asset: Asset, llm: MultimodalClient, data_root: Path) -> ProcessingResult:
    src = data_root / asset.storage_path
    work = data_root / "_work" / str(asset.id)
    result = ProcessingResult()

    try:
        result.duration = _media_duration(src)

        thumb = _first_frame(src, data_root / "thumbnails" / f"{asset.id}.jpg")
        if thumb:
            result.thumbnail_path = f"thumbnails/{thumb.name}"

        frames = _extract_keyframes(src, work, settings.video_max_frames, result.duration)
        descs: list[str] = []
        tags: list[str] = []
        ocrs: list[str] = []
        for f in frames:
            try:
                with Image.open(f) as img:
                    img.load()
                    model = _route_vision_model(max(img.size))
                    vision = llm.vision_describe(_image_to_b64(img), "image/jpeg", model=model)
                    result.vision_model = model
            except Exception as e:
                logger.warning("关键帧理解失败 %s: %s", f.name, e)
                continue
            if vision:
                record_usage(asset.id, model, "vision")
                if vision.description:
                    descs.append(vision.description)
                tags.extend(vision.tags)
                if vision.ocr:
                    ocrs.append(vision.ocr)
        result.description = _dedupe_join(descs)
        result.tags = _top_tags(tags)
        result.ocr_text = _dedupe_join(ocrs, sep="\n", limit=2000)

        wav = _extract_audio(src, work / "audio.wav", settings.audio_max_seconds)
        if wav:
            transcript, segments = _transcribe_segments(asset, llm, wav, work, result.duration)
            if transcript:
                result.transcript = transcript
                result.transcript_segments = json.dumps(segments, ensure_ascii=False) if segments else None

        if not result.description and result.transcript:
            summary = llm.summarize_text(result.transcript[:3000])
            if summary:
                record_usage(asset.id, settings.llm_model, "summary")
                result.description = summary.summary
                result.tags = summary.tags
    finally:
        shutil.rmtree(work, ignore_errors=True)
        try:
            work.parent.rmdir()  # 若 _work 已空则顺手删除
        except OSError:
            pass

    if not result.description and not result.transcript:
        result.description = "（视频理解未启用或失败，仅保留封面与元数据）"
    return result


# ---------- 音频 ----------

def process_audio(asset: Asset, llm: MultimodalClient, data_root: Path) -> ProcessingResult:
    src = data_root / asset.storage_path
    work = data_root / "_work" / str(asset.id)
    result = ProcessingResult()

    try:
        result.duration = _media_duration(src)

        thumb = _audio_placeholder(data_root / "thumbnails" / f"{asset.id}.jpg")
        result.thumbnail_path = f"thumbnails/{thumb.name}"

        wav = _extract_audio(src, work / "audio.wav", settings.audio_max_seconds)
        if wav:
            transcript, segments = _transcribe_segments(asset, llm, wav, work, result.duration)
            if transcript:
                result.transcript = transcript
                result.transcript_segments = json.dumps(segments, ensure_ascii=False) if segments else None

        if result.transcript:
            summary = llm.summarize_text(result.transcript[:3000])
            if summary:
                record_usage(asset.id, settings.llm_model, "summary")
                result.description = summary.summary
                result.tags = summary.tags
    finally:
        shutil.rmtree(work, ignore_errors=True)
        try:
            work.parent.rmdir()
        except OSError:
            pass

    if not result.description:
        result.description = "（音频转写未启用或失败，仅保留元数据）"
    return result


_PROCESSORS = {
    "image": process_image,
    "document": process_document,
    "video": process_video,
    "audio": process_audio,
}


def resolve_processor(modality: str):
    return _PROCESSORS[modality]


def build_embed_text(result: ProcessingResult) -> str:
    parts = [result.description, result.ocr_text, result.transcript, result.text_content]
    return " ".join(p for p in parts if p)[:1500]
